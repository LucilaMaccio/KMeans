import json
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook

# --------------------------------------------------------------------------------------
# POC
# --------------------------------------------------------------------------------------

# Open the workbook and select a worksheet
# wb = load_workbook("C:/Users/lucil/OneDrive/EAE/TFM/poc_xlsx_to_json.xlsx")

# sheet = wb['Cars']

# # List to hold dictionaries
# cars_list = []

# # Iterate through each row in worksheet and fetch values into dict
# for row in islice(sheet.values, 1, sheet.max_row):
#     cars = OrderedDict()
#     cars['car-id'] = row[0]
#     cars['maker'] = row[1]
#     cars['model'] = row[2]
#     cars['miles'] = row[3]

#     print(cars)

#     cars_list.append(cars)

# # Serialize the list of dicts to JSON
# j = json.dumps(cars_list)

# # Write to file
# with open('data.json', 'w') as f:
#     f.write(j)


# --------------------------------------------------------------------------------------
# PHARMACIES
# --------------------------------------------------------------------------------------

# wb = load_workbook("C:/Users/lucil/OneDrive/EAE/TFM/datasets/xlsx/turnosmadrid_guardia_2021.xlsx")

# sheet = wb['turnosmadrid_guardia_2021']

# # List to hold dictionaries
# pharmacies_list = []

# # Iterate through each row in worksheet and fetch values into dict
# for row in islice(sheet.values, 1, sheet.max_row):
#     pharmacies = OrderedDict()
#     pharmacies['identifier'] = row[3]
#     pharmacies['location'] = row[1]
#     pharmacies['address'] = row[4]
#     pharmacies['service_type'] = row[5]

#     # print(pharmacies)

#     alreadyExists = False

#     if len(pharmacies_list) == 0:
#         pharmacies_list.append(pharmacies)
#     else:
#         for pharmacy in pharmacies_list:
#             if pharmacy['identifier'] == pharmacies['identifier']:
#                 alreadyExists = True
        
#         if not alreadyExists:
#             pharmacies_list.append(pharmacies)

# print(len(pharmacies_list))

# # Serialize the list of dicts to JSON
# j = json.dumps(pharmacies_list, ensure_ascii=False)

# # Write to file
# with open('C:/Users/lucil/OneDrive/EAE/TFM/datasets/json/pharmacies.json', 'w', encoding='utf-8') as f:
#     f.write(j)

# --------------------------------------------------------------------------------------
# GREEN ZONES
# --------------------------------------------------------------------------------------

wb = load_workbook("C:/Users/lucil/OneDrive/EAE/TFM/datasets/xlsx/IZVER_ZONAS_VERDES_2019.xlsx")

sheet = wb['IZVERD_2019_CLEANED']

# List to hold dictionaries
greenZones_list = []

# Iterate through each row in worksheet and fetch values into dict
for row in islice(sheet.values, 1, sheet.max_row):
    greenZones = OrderedDict()
    greenZones['identifier'] = row[3]
    greenZones['location'] = row[0]
    greenZones['address'] = row[5]
    greenZones['size'] = row[7]

    # print(greenZones)

    greenZones_list.append(greenZones)

print(len(greenZones_list))

# Serialize the list of dicts to JSON
j = json.dumps(greenZones_list, ensure_ascii=False)

# Write to file
with open('C:/Users/lucil/OneDrive/EAE/TFM/datasets/json/green_zones.json', 'w', encoding='utf-8') as f:
    f.write(j)

# --------------------------------------------------------------------------------------
# SECURITY
# --------------------------------------------------------------------------------------

# wb = load_workbook("C:/Users/lucil/OneDrive/EAE/TFM/datasets/xlsx/Datos_Marzo_2021-copia.xlsx")

# securitySheet = wb['SEGURIDAD']
# accidentsSheet = wb['ACCIDENTES']

# # List to hold dictionaries
# security_list = []
# accidents_list = []

# # Iterate through each row in worksheet and fetch values into dict
# for row in islice(securitySheet.values, 1, 22):
#     security = OrderedDict()
#     security['location'] = row[0]
#     security['security_related_with_people'] = row[1]
#     security['security_related_with_patrimony'] = row[2]
#     security_list.append(security)

# for row in islice(accidentsSheet.values, 1, 22):
#     accidents = OrderedDict()
#     accidents['location'] = row[0]
#     accidents['accidents_with_victims'] = row[1]
#     accidents['accidents_without_victims'] = row[2]
#     accidents_list.append(accidents)
    
# print(len(security_list))
# print(len(accidents_list))

# # Serialize the list of dicts to JSON
# securityJson = json.dumps(security_list, ensure_ascii=False)
# accidentsJson = json.dumps(accidents_list, ensure_ascii=False)

# # Write to file
# with open('C:/Users/lucil/OneDrive/EAE/TFM/datasets/json/security.json', 'w', encoding='utf-8') as f1:
#     f1.write(securityJson)

# with open('C:/Users/lucil/OneDrive/EAE/TFM/datasets/json/accidents.json', 'w', encoding='utf-8') as f2:
#     f2.write(accidentsJson)
    
# --------------------------------------------------------------------------------------
# LIBRARIES
# --------------------------------------------------------------------------------------

# wb = load_workbook("C:/Users/lucil/OneDrive/EAE/TFM/datasets/xlsx/201747-0-bibliobuses-bibliotecas.xlsx")

# sheet = wb['201747-0-bibliobuses-biblioteca']

# # List to hold dictionaries
# libraries_list = []

# # Iterate through each row in worksheet and fetch values into dict
# for row in islice(sheet.values, 1, sheet.max_row):
#     libraries = OrderedDict()
#     libraries['identifier'] = row[0]
#     libraries['location'] = row[21]
#     libraries_list.append(libraries)
    
# print(len(libraries_list))

# # Serialize the list of dicts to JSON
# j = json.dumps(libraries_list, ensure_ascii=False)

# # Write to file
# with open('C:/Users/lucil/OneDrive/EAE/TFM/datasets/json/libraries.json', 'w', encoding='utf-8') as f:
#     f.write(j)

# --------------------------------------------------------------------------------------
# MIDDLE SCHOOLS
# --------------------------------------------------------------------------------------

# wb = load_workbook("C:/Users/lucil/OneDrive/EAE/TFM/datasets/xlsx/202311-0-colegios-publicos.xlsx")

# sheet = wb['202311-0-colegios-publicos']

# # List to hold dictionaries
# schools_list = []

# # Iterate through each row in worksheet and fetch values into dict
# for row in islice(sheet.values, 1, sheet.max_row):
#     schools = OrderedDict()
#     schools['identifier'] = row[0]
#     schools['location'] = row[21]
#     schools_list.append(schools)
    
# print(len(schools_list))

# # Serialize the list of dicts to JSON
# j = json.dumps(schools_list, ensure_ascii=False)

# # Write to file
# with open('C:/Users/lucil/OneDrive/EAE/TFM/datasets/json/middle_schools.json', 'w', encoding='utf-8') as f:
#     f.write(j)

# --------------------------------------------------------------------------------------
# SOCIAL ATTENTION - FAMILIES
# --------------------------------------------------------------------------------------

# wb = load_workbook("C:/Users/lucil/OneDrive/EAE/TFM/datasets/xlsx/205244-0-infancia-familia-adolescentes.xlsx")

# sheet = wb['205244-0-infancia-familia-adole']

# # List to hold dictionaries
# social_attention_list = []

# # Iterate through each row in worksheet and fetch values into dict
# for row in islice(sheet.values, 1, sheet.max_row):
#     social_attention = OrderedDict()
#     social_attention['identifier'] = row[0]
#     social_attention['name'] = row[1]
#     social_attention['location'] = row[21]
#     social_attention_list.append(social_attention)
    
# print(len(social_attention_list))

# # Serialize the list of dicts to JSON
# j = json.dumps(social_attention_list, ensure_ascii=False)

# # Write to file
# with open('C:/Users/lucil/OneDrive/EAE/TFM/datasets/json/social_attention.json', 'w', encoding='utf-8') as f:
#     f.write(j)

# --------------------------------------------------------------------------------------
# SPORT CENTERS
# --------------------------------------------------------------------------------------

# wb = load_workbook("C:/Users/lucil/OneDrive/EAE/TFM/datasets/xlsx/200215-0-instalaciones-deportivas.xlsx")

# sheet = wb['200215-0-instalaciones-deportiv']

# # List to hold dictionaries
# sport_centers_list = []

# # Iterate through each row in worksheet and fetch values into dict
# for row in islice(sheet.values, 1, sheet.max_row):
#     sports = OrderedDict()
#     sports['identifier'] = row[0]
#     sports['location'] = row[21]
#     sport_centers_list.append(sports)
    
# print(len(sport_centers_list))

# # Serialize the list of dicts to JSON
# j = json.dumps(sport_centers_list, ensure_ascii=False)

# # Write to file
# with open('C:/Users/lucil/OneDrive/EAE/TFM/datasets/json/sport_centers.json', 'w', encoding='utf-8') as f:
#     f.write(j)

# --------------------------------------------------------------------------------------
# CATHOLIC CHURCHES
# --------------------------------------------------------------------------------------

# wb = load_workbook("C:/Users/lucil/OneDrive/EAE/TFM/datasets/xlsx/209426-0-templos-catolicas.xlsx")

# sheet = wb['209426-0-templos-catolicas']

# # List to hold dictionaries
# cath_churches_list = []

# # Iterate through each row in worksheet and fetch values into dict
# for row in islice(sheet.values, 1, sheet.max_row):
#     cath_churches = OrderedDict()
#     cath_churches['identifier'] = row[0]
#     cath_churches['name'] = row[1]
#     cath_churches['location'] = row[21]
#     cath_churches_list.append(cath_churches)
    
# print(len(cath_churches_list))

# # Serialize the list of dicts to JSON
# j = json.dumps(cath_churches_list, ensure_ascii=False)

# # Write to file
# with open('C:/Users/lucil/OneDrive/EAE/TFM/datasets/json/cath_churches.json', 'w', encoding='utf-8') as f:
#     f.write(j)

# --------------------------------------------------------------------------------------
# NON CATHOLIC CHURCHES
# --------------------------------------------------------------------------------------

# wb = load_workbook("C:/Users/lucil/OneDrive/EAE/TFM/datasets/xlsx/209434-0-templos-otros.xlsx")

# sheet = wb['209434-0-templos-otros']

# # List to hold dictionaries
# non_cath_churches_list = []

# # Iterate through each row in worksheet and fetch values into dict
# for row in islice(sheet.values, 1, sheet.max_row):
#     non_cath_churches = OrderedDict()
#     non_cath_churches['identifier'] = row[0]
#     non_cath_churches['name'] = row[1]
#     non_cath_churches['location'] = row[21]
#     non_cath_churches_list.append(non_cath_churches)
    
# print(len(non_cath_churches_list))

# # Serialize the list of dicts to JSON
# j = json.dumps(non_cath_churches_list, ensure_ascii=False)

# # Write to file
# with open('C:/Users/lucil/OneDrive/EAE/TFM/datasets/json/non_cath_churches.json', 'w', encoding='utf-8') as f:
#     f.write(j)

# --------------------------------------------------------------------------------------
# MARKETS
# --------------------------------------------------------------------------------------

# wb = load_workbook("C:/Users/lucil/OneDrive/EAE/TFM/datasets/xlsx/200967-0-mercados.xlsx")

# sheet = wb['200967-0-mercados']

# # List to hold dictionaries
# markets_list = []

# # Iterate through each row in worksheet and fetch values into dict
# for row in islice(sheet.values, 1, sheet.max_row):
#     markets = OrderedDict()
#     markets['identifier'] = row[0]
#     markets['name'] = row[1]
#     markets['location'] = row[21]
#     markets_list.append(markets)
    
# print(len(markets_list))

# # Serialize the list of dicts to JSON
# j = json.dumps(markets_list, ensure_ascii=False)

# # Write to file
# with open('C:/Users/lucil/OneDrive/EAE/TFM/datasets/json/markets.json', 'w', encoding='utf-8') as f:
#     f.write(j)

# --------------------------------------------------------------------------------------
# POOLS
# --------------------------------------------------------------------------------------

# wb = load_workbook("C:/Users/lucil/OneDrive/EAE/TFM/datasets/xlsx/210227-0-piscinas-publicas.xlsx")

# sheet = wb['210227-0-piscinas-publicas']

# # List to hold dictionaries
# pools_list = []

# # Iterate through each row in worksheet and fetch values into dict
# for row in islice(sheet.values, 1, sheet.max_row):
#     pools = OrderedDict()
#     pools['identifier'] = row[0]
#     pools['name'] = row[1]
#     pools['location'] = row[21]
#     pools_list.append(pools)
    
# print(len(pools_list))

# # Serialize the list of dicts to JSON
# j = json.dumps(pools_list, ensure_ascii=False)

# # Write to file
# with open('C:/Users/lucil/OneDrive/EAE/TFM/datasets/json/pools.json', 'w', encoding='utf-8') as f:
#     f.write(j)